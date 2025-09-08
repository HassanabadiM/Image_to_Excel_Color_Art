from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

image = Image.open('sample_image.png')

if image.mode == 'RGBA':
    image = image.convert('RGB')

width, height = image.size
result_width = width // 10
result_height = height // 10

wb = Workbook()
ws = wb.active
ws.title = "Average Colors"

for y in range(result_height):
    for x in range(result_width):
        r_total, g_total, b_total, count = 0, 0, 0, 0
        
        for dy in range(10):
            for dx in range(10):
                px = x * 10 + dx
                py = y * 10 + dy
                if px < width and py < height:
                    pixel = image.getpixel((px, py))
                    r, g, b = pixel[0], pixel[1], pixel[2]
                    r_total += r
                    g_total += g
                    b_total += b
                    count += 1  

        avg_r = r_total // count
        avg_g = g_total // count
        avg_b = b_total // count
        
        hex_color = f"{avg_r:02X}{avg_g:02X}{avg_b:02X}"
            
        cell = ws.cell(row=y+1, column=x+1)
        cell.value = f"#{hex_color}"
            
        cell.fill = PatternFill(start_color=hex_color, 
                                end_color=hex_color, 
                                fill_type="solid")


for row in range(1, result_height + 1):
    ws.row_dimensions[row].height = 20

for col in range(1, result_width + 1):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = 3

wb.save("average_colors.xlsx")
print('Saved Successfully!')
